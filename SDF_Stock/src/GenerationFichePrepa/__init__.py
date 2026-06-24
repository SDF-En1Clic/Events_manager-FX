import logging
import azure.functions as func
from azure.identity import DefaultAzureCredential
from azure.keyvault.secrets import SecretClient
import requests
import urllib.parse
import json
from datetime import datetime, timezone
import io
import re
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment

try:
    from zoneinfo import ZoneInfo
    LOCAL_TZ = ZoneInfo("Europe/Paris")
except Exception:  # tzdata absent
    LOCAL_TZ = None

# --------- CONFIG GLOBALE -------------
VAULT_URL = "https://events-manager-kv.vault.azure.net/"
TEMPLATE_PATH = "00 - Templates/Template_fiche_prepa_V1.xlsx"
TARGET_FOLDER = "03 - Documents commandes"
TYPE_DOC = "Fiche préparation"
# --------------------------------------

session = requests.Session()
adapter = requests.adapters.HTTPAdapter(pool_connections=100, pool_maxsize=100)
session.mount('https://', adapter)


# =========================================================================
#  KEY VAULT / GRAPH HELPERS  (repris de GenerationDocument)
# =========================================================================
def get_secret(name: str):
    credential = DefaultAzureCredential()
    client = SecretClient(vault_url=VAULT_URL, credential=credential)
    return client.get_secret(name).value


def get_graph_token(tenant_id, client_id, client_secret):
    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default"
    }
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    try:
        response = session.post(url, data=data, headers=headers)
        response.raise_for_status()
        return response.json().get("access_token")
    except Exception as e:
        logging.error(f"Erreur token: {e}")
        return None


def graph_filtered_items(site_id, list_id, token, filter_expr=None):
    base_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items?$expand=fields"
    headers = {
        "Authorization": f"Bearer {token}",
        "Prefer": "HonorNonIndexedQueriesWarningMayFailRandomly"
    }
    if filter_expr:
        filter_param = urllib.parse.quote(filter_expr, safe="=()/ '")
        base_url += f"&$filter={filter_param}"

    results = []
    url = base_url
    while url:
        res = session.get(url, headers=headers)
        if not res.ok:
            logging.error(f"Erreur API Graph (filtré). Status: {res.status_code}. Réponse: {res.text}")
        res.raise_for_status()
        data = res.json()
        results.extend(data.get("value", []))
        url = data.get("@odata.nextLink")

    out = []
    for item in results:
        fields = item.get("fields", {})
        # On conserve l'id Graph de l'élément (= ID de liste) pour les PatchItem.
        fields.setdefault("_item_id", item.get("id"))
        out.append(fields)
    return out


def graph_get_item_by_id(site_id, list_id, item_id, token):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items/{item_id}?$expand=fields"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        res = session.get(url, headers=headers)
        res.raise_for_status()
        return res.json().get("fields", {})
    except Exception as e:
        logging.error(f"Erreur Get Item: {e}")
        return None


def download_graph_file(site_id, token, filepath):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    headers = {"Authorization": f"Bearer {token}"}
    res = session.get(url, headers=headers)
    res.raise_for_status()
    drives = res.json().get("value", [])

    drive_id = None
    for d in drives:
        if d.get("name") in ("Documents partages", "Documents partagés"):
            drive_id = d.get("id")
            break
    if not drive_id and drives:
        drive_id = drives[0].get("id")

    url_file = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{urllib.parse.quote(filepath)}:/content"
    res_file = session.get(url_file, headers=headers)
    res_file.raise_for_status()
    return res_file.content


def graph_post_item(site_id, list_id, token, fields_dict):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    res = session.post(url, headers=headers, json={"fields": fields_dict})
    if not res.ok:
        logging.error(f"Erreur create item: {res.status_code} {res.text}")
    res.raise_for_status()
    return res.json().get("id")


def graph_patch_item(site_id, list_id, item_id, token, fields_dict):
    """Met à jour les champs d'un élément de liste (PatchItem du flux)."""
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items/{item_id}/fields"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        res = session.patch(url, headers=headers, json=fields_dict)
        if not res.ok:
            logging.error(f"Erreur patch item {item_id}: {res.status_code} {res.text}")
        return res.ok
    except Exception as e:
        logging.error(f"Exception patch item {item_id}: {e}")
        return False


# =========================================================================
#  EXCEL HELPERS
# =========================================================================
# Cellules d'en-tête (rows 1-11), identiques sur les 4 onglets de vue.
HEADER_CELLS = {
    "A2": "Emetteur", "E2": "Destinataire",
    "B3": "Commande N°", "E3": "CREATEUR",
    "B4": "Client",
    "B6": "Adresse Livraison", "B7": "Ville Livraison",
    "B8": "Site de livraison", "B9": "Telephone",
    "B10": "Date de livraison", "B11": "Date de tir",
}
HEADER_ROW = 12  # ligne des libellés de colonnes ; les données commencent en 13


def fill_view_sheet(ws, header_data, data_rows, col_keys):
    """Remplit un onglet de vue en VALEURS STATIQUES (en-tête + données).

    On écrase toutes les formules matricielles du template (en-tête + zone de
    données) : Excel n'a donc plus de tableau dynamique « spill » à réparer à
    l'ouverture, et le contenu est entièrement maîtrisé.
    """
    # En-tête
    for cell, key in HEADER_CELLS.items():
        ws[cell] = header_data.get(key, "")

    # Données (à partir de la ligne 13). On réplique le style de la ligne modèle
    # (1ʳᵉ ligne de données du template) sur toutes les lignes écrites pour un
    # rendu homogène, et on remet à « Normal » les lignes vides.
    first = HEADER_ROW + 1
    n_cols = max(len(col_keys), ws.max_column)
    model_styles = [copy(ws.cell(row=first, column=c)._style) for c in range(1, n_cols + 1)]
    model_height = ws.row_dimensions[first].height

    rows_to_clear = max(len(data_rows), ws.max_row - HEADER_ROW)
    for i in range(rows_to_clear):
        r = first + i
        values = data_rows[i] if i < len(data_rows) else None
        has_data = values is not None
        for c in range(1, n_cols + 1):
            key = col_keys[c - 1] if c <= len(col_keys) else None
            cell = ws.cell(row=r, column=c)
            cell.value = values.get(key, "") if (has_data and key is not None) else None
            if has_data:
                cell._style = copy(model_styles[c - 1])
                # Alignement : REFERENCE à gauche, tout le reste centré
                al = cell.alignment
                horiz = "left" if key == "REFERENCE" else "center"
                cell.alignment = Alignment(horizontal=horiz, vertical=al.vertical or "center",
                                           wrap_text=al.wrap_text)
            else:
                cell.style = "Normal"
        ws.row_dimensions[r].height = model_height if has_data else None


_EXTERNAL_REF = re.compile(r'\[\d+\]')


def _strip_external_defined_names(wb):
    """Supprime les noms définis qui référencent un classeur externe ([n]…) ou #REF.

    (Hérités du .xlsm source ; ils deviennent invalides une fois le lien externe
    retiré et provoquent la « réparation » d'Excel à l'ouverture.)
    """
    def _drop(container):
        try:
            for nm in list(container):
                val = getattr(container[nm], "value", "") or ""
                if _EXTERNAL_REF.search(val) or "#REF" in val:
                    del container[nm]
        except Exception as e:
            logging.warning(f"Nettoyage noms définis: {e}")

    _drop(wb.defined_names)
    for ws in wb.worksheets:
        _drop(ws.defined_names)


def find_sheet(wb, name):
    """Retrouve une feuille par nom exact, sinon par préfixe (accents/encodage)."""
    if name in wb.sheetnames:
        return wb[name]
    pref = name[:5].lower()
    for ws in wb.worksheets:
        if ws.title.lower().startswith(pref):
            return ws
    return None


# =========================================================================
#  UTILITAIRES MÉTIER
# =========================================================================
def clean_id(val):
    if val is None or val == "":
        return ""
    if isinstance(val, float):
        return str(int(val))
    return str(val)


def fmt_date(val):
    """Formate une date SharePoint (UTC) en jj/mm/aaaa, heure locale Europe/Paris.

    SharePoint renvoie les dates en UTC (ex. '2026-06-17T22:00:00Z' = 18/06 à Paris).
    On convertit donc en fuseau local avant d'extraire la date, pour éviter le
    décalage d'un jour.
    """
    if not val:
        return ""
    s = str(val)
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        if LOCAL_TZ is not None:
            dt = dt.astimezone(LOCAL_TZ)
        return dt.strftime("%d/%m/%Y")
    except Exception:
        pass
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return val


def fmt_qte(val):
    """Quantité lisible (entier si possible)."""
    if val is None or val == "":
        return ""
    try:
        f = float(val)
        return int(f) if f.is_integer() else f
    except Exception:
        return val


def fetch_details(site_id, list_id, token, commande_id):
    """Détails commande, en gérant CMD_ID texte ou numérique (comme GenerationDocument)."""
    details = []
    try:
        details = graph_filtered_items(site_id, list_id, token, f"fields/CMD_ID eq '{commande_id}'")
    except Exception:
        pass
    if not details:
        try:
            details = graph_filtered_items(site_id, list_id, token, f"fields/CMD_ID eq {commande_id}")
        except Exception:
            pass
    return details


def tri_lignes(item):
    val = str(item.get("Title", "")).replace("Ligne", "").strip()
    try:
        return (0, float(val), str(item.get("Reference", "")).strip().upper())
    except ValueError:
        return (1, val, str(item.get("Reference", "")).strip().upper())


def ligne_num(r):
    """Clé de tri numérique sur le numéro de ligne ('Ligne 12' -> 12)."""
    val = str(r.get("LIGNE", "")).replace("Ligne", "").strip()
    try:
        return (0, float(val))
    except ValueError:
        return (1, val)


def is_accessoire(prod):
    """Un accessoire = Famille_FWSIM == 'ACCESSOIRE' ET Sous_famille_FWSIM != 'UKOBA'.

    (Logique PowerApps : ces lignes vont dans l'onglet Accessoires ; tout le reste,
    y compris les accessoires de sous-famille UKOBA, va dans Fiche_Prepa.)
    """
    fam = str(prod.get("Famille_FWSIM", "")).strip().upper()
    sous = str(prod.get("Sous_famille_FWSIM", "")).strip().upper()
    return fam == "ACCESSOIRE" and sous != "UKOBA"


# =========================================================================
#  CONSTRUCTION DES DONNÉES D'UN FICHIER (pour un site de stock)
# =========================================================================
def build_fiche_prepa_rows(details_site, produits_map, inventaire_cache,
                           site_id, inventaire_list_id, details_list_id, token,
                           site_filter, is_secondary):
    """Construit les lignes de l'onglet Fiche_Prepa pour un site de stock
    et applique le PatchItem sur la liste Commandes_details.

    Contenu = tous les produits NON-accessoires (les accessoires non-UKOBA vont
    dans l'onglet Accessoires). inventaire_cache : {(ref, scope): {...}} pour
    éviter les requêtes répétées. site_filter : valeur de Site_Stock_second.
    """
    rows = []
    for det in details_site:
        ref = str(det.get("Reference", "")).strip().upper()
        prod = produits_map.get(ref, {})
        # On ne garde que les produits dont l'origine n'est pas UKOBA
        if str(prod.get("Origine", "")).strip().upper() == "UKOBA":
            continue
        # Onglet principal = tout sauf les accessoires (non-UKOBA)
        if is_accessoire(prod):
            continue

        # --- Emplacement / Bâtiment depuis l'inventaire ---
        cache_key = (ref, "sec" if is_secondary else "prim")
        if cache_key not in inventaire_cache:
            if is_secondary:
                inv_filter = f"fields/Title eq '{ref}' and fields/Site eq '{site_filter}'"
            else:
                inv_filter = f"fields/Title eq '{ref}' and fields/Site ne '{site_filter}'"
            bat, empl = "", ""
            try:
                inv_items = graph_filtered_items(site_id, inventaire_list_id, token, inv_filter)
                if inv_items:
                    bat = inv_items[0].get("Batiment", "") or ""
                    empl = inv_items[0].get("Emplacement", "") or ""
            except Exception as e:
                logging.warning(f"Inventaire introuvable pour {ref}: {e}")
            inventaire_cache[cache_key] = {"bat": bat, "empl": empl}

        inv = inventaire_cache[cache_key]
        bat_prefix = "P" if det.get("Emplacement") else ""
        designation = prod.get("Description_pyromotion", "") or det.get("Description_pyro", "")

        rows.append({
            "LIGNE": f"Ligne {det.get('Title', '')}",
            "REFERENCE": ref,
            "DESIGNATION": designation,
            "QT": fmt_qte(det.get("Quantite", "")),
            "COMMENTAIRE": det.get("Commentaires", ""),
            "SITE": det.get("Site", ""),
            "EMPLACEMENT": f"{bat_prefix}{inv['empl']}",
            "STATUT": det.get("Statut", ""),
            "ORIGINE": prod.get("Origine", ""),
            "BATIMENT": inv["bat"],
            "CASE": "☐",
        })

        # --- PatchItem -> Commandes_details (PAS l'inventaire) ---
        # Uniquement si l'inventaire a renvoyé un emplacement/bâtiment, pour ne
        # pas écraser des données existantes par du vide.
        det_id = det.get("_item_id") or det.get("ID") or det.get("id")
        if det_id and (inv["empl"] or inv["bat"]):
            graph_patch_item(site_id, details_list_id, det_id, token, {
                "Emplacement": inv["empl"],
                "Batiment": inv["bat"],
                "Comptabilise_inventaire": 0,
            })

    # Onglet Fiche_Prepa : trié par numéro de ligne
    rows.sort(key=ligne_num)
    return rows


def build_grouped_rows(tab_datas_rows):
    """Onglet Classés par produit : trié par RÉFÉRENCE puis groupé par RÉFÉRENCE
    (en-tête de groupe + lignes enfants triées par ligne)."""
    rows_sorted = sorted(tab_datas_rows, key=lambda r: (str(r.get("REFERENCE", "")), ligne_num(r)))
    grouped = []
    prev_ref = None
    for r in rows_sorted:
        cur_ref = r.get("REFERENCE", "")
        if cur_ref != prev_ref:
            grouped.append({"REFERENCE": cur_ref})  # ligne d'en-tête de groupe (pas de case)
        grouped.append({
            "LIGNE": r.get("LIGNE", ""),
            "REFERENCE": "",
            "DESIGNATION": r.get("DESIGNATION", ""),
            "QT": r.get("QT", ""),
            "EMPLACEMENT": r.get("EMPLACEMENT", ""),
            "STATUT": r.get("STATUT", ""),
            "ORIGINE": r.get("ORIGINE", ""),
            "BATIMENT": r.get("BATIMENT", ""),
            "CASE": "☐",
        })
        prev_ref = cur_ref
    return grouped


def build_global_rows(fiche_rows, accessoires_rows):
    """Onglet Global : produits + accessoires (tout sauf UKOBA), groupé par référence."""
    acc = [{
        "LIGNE": "",
        "REFERENCE": a.get("REFERENCE", ""),
        "DESIGNATION": a.get("DESIGNATION", ""),
        "QT": a.get("QTE", ""),
        "ORIGINE": a.get("ORIGINE", ""),
        "STATUT": a.get("STATUT", ""),
        "BATIMENT": a.get("BATIMENT", ""),
        "EMPLACEMENT": a.get("EMPLACEMENT", ""),
    } for a in accessoires_rows]
    return build_grouped_rows(fiche_rows + acc)


def build_accessoires_rows(details_site, produits_map):
    """Onglet Accessoires : détails dont la référence est un accessoire (non-UKOBA)."""
    rows = []
    for det in details_site:
        ref = str(det.get("Reference", "")).strip().upper()
        prod = produits_map.get(ref, {})
        # On ne garde que les produits dont l'origine n'est pas UKOBA
        if str(prod.get("Origine", "")).strip().upper() == "UKOBA":
            continue
        if not is_accessoire(prod):
            continue
        if not prod.get("Description_pyromotion"):
            continue
        rows.append({
            "REFERENCE": ref,
            "DESIGNATION": prod.get("Description_pyromotion", ""),
            "ORIGINE": prod.get("Origine", ""),
            "QTE": fmt_qte(det.get("Quantite", "")),
            "STATUT": det.get("Statut", ""),
            "BATIMENT": det.get("Batiment", ""),
            "EMPLACEMENT": det.get("Emplacement", ""),
            "CASE": "☐",
        })
    rows.sort(key=lambda r: str(r.get("REFERENCE", "")))
    return rows


def build_materiel_rows(reservations_site, materiel_map):
    """Onglet Materiel : réservations (par Aff_ID) jointes à la liste Materiel.

    La jointure se fait sur le Materiel_ID qui correspond au champ interne
    `Title` dans les deux listes (affiché « Materiel_ID » dans SharePoint).
    Batiment/Emplacement viennent de la liste Materiel (absents de la réservation).
    """
    rows = []
    for resa in reservations_site:
        mat_id = clean_id(resa.get("Title"))
        mat = materiel_map.get(mat_id, {})
        reference = mat_id
        rows.append({
            "REFERENCE": reference,
            "DESIGNATION": mat.get("Designation", ""),
            "CATEGORIE": mat.get("Categorie", ""),
            "QTE": fmt_qte(resa.get("Quantite", "")),
            "STATUT": resa.get("Statut", "") or mat.get("Statut", ""),
            "BATIMENT": resa.get("Batiment", "") or mat.get("Batiment", ""),
            "EMPLACEMENT": resa.get("Emplacement", "") or mat.get("Emplacement", ""),
            "CASE": "☐",
        })
    rows.sort(key=lambda r: str(r.get("REFERENCE", "")))
    return rows


def fill_workbook(template_bytes, header_data, fiche_rows, grouped_rows, accessoires_rows, materiel_rows):
    """Remplit les 4 onglets de vue en valeurs statiques (un fichier = un site de stock).

    Aucune formule conservée : pas de tableau dynamique « spill » que Excel
    devrait réparer à l'ouverture. L'ordre des colonnes suit les libellés des
    onglets du template.
    """
    wb = load_workbook(io.BytesIO(template_bytes))

    # Nettoyage des éléments inutiles à une sortie statique et sources d'erreurs
    # de « réparation » à l'ouverture Excel : tables, liens externes (le template
    # hérite d'un lien vers un classeur local), métadonnées de tableaux dynamiques,
    # et les noms définis qui pointaient vers ce classeur externe (sinon #REF).
    for ws in wb.worksheets:
        for tname in list(ws.tables):
            del ws.tables[tname]
    try:
        wb._external_links = []
    except Exception:
        pass
    _strip_external_defined_names(wb)

    # Chaque onglet est cherché par son nom cible (1er) PUIS son ancien nom, pour
    # fonctionner que le template ait déjà été renommé (Produits/Global) ou non
    # (Fiche_Prepa/Classés par produit). Le 1er nom est le nom final appliqué.
    layouts = [
        (("Produits", "Fiche_Prepa"), fiche_rows,
         ["LIGNE", "REFERENCE", "DESIGNATION", "QT", "ORIGINE", "STATUT", "BATIMENT", "EMPLACEMENT", "CASE"]),
        (("Global", "Classés par produit"), grouped_rows,
         ["REFERENCE", "LIGNE", "DESIGNATION", "QT", "ORIGINE", "STATUT", "BATIMENT", "EMPLACEMENT", "CASE"]),
        (("Accessoires",), accessoires_rows,
         ["REFERENCE", "DESIGNATION", "ORIGINE", "QTE", "STATUT", "BATIMENT", "EMPLACEMENT", "CASE"]),
        (("Materiel",), materiel_rows,
         ["REFERENCE", "DESIGNATION", "CATEGORIE", "QTE", "STATUT", "BATIMENT", "EMPLACEMENT", "CASE"]),
    ]
    for names, rows, cols in layouts:
        ws = None
        for nm in names:
            ws = find_sheet(wb, nm)
            if ws is not None:
                break
        if ws is not None:
            fill_view_sheet(ws, header_data, rows, cols)
            if ws.title != names[0]:
                ws.title = names[0]  # applique le nom final

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================================================================
#  POINT D'ENTRÉE
# =========================================================================
def main(req: func.HttpRequest) -> func.HttpResponse:
    try:
        body = req.get_json()
        commande_id = body.get("ID_cmd")
        type_doc = body.get("Type_Doc", TYPE_DOC)

        logging.info(f"Paramètres reçus : ID_cmd={commande_id}, Type_Doc={type_doc}")
        if not commande_id:
            return _err("Paramètre 'ID_cmd' requis", 400)
        if type_doc != TYPE_DOC:
            return _err(f"Type_Doc non supporté: {type_doc}", 400)

        # --- Secrets ---
        tenant_id = get_secret("tenantid")
        client_id = get_secret("clientid")
        client_secret = get_secret("appsecret")
        site_id = get_secret("siteid")

        commandes_list_id = get_secret("cmdlistid")
        details_list_id = get_secret("cmddetailslistid")
        produits_list_id = get_secret("produitslistid")
        admin_list_id = get_secret("configlistid")
        affaire_list_id = get_secret("affaireevtslistid")
        clients_list_id = get_secret("clientslistid")
        commande_doc_list_id = get_secret("commandedoclistid")
        inventaire_list_id = get_secret("inventairelistid")
        materiel_list_id = get_secret("materiellistid")
        materiel_resa_list_id = get_secret("materielreservationlistid")

        token = get_graph_token(tenant_id, client_id, client_secret)
        if not token:
            return _err("Échec de l'authentification Graph", 500)

        # --- 1. Commande ---
        commande = graph_get_item_by_id(site_id, commandes_list_id, commande_id, token)
        if not commande:
            return _err("Commande introuvable", 404)

        entite = commande.get("Entite", "")
        numero_commande = commande.get("Title", "")
        site_livraison_title = commande.get("Site_livraison", "")
        date_livraison = fmt_date(commande.get("Date_livraison", ""))
        site_stock = commande.get("Site_Stock", "")
        site_stock_second = commande.get("Site_Stock_second", "")

        # --- 2. Affaire / Client ---
        id_aff = clean_id(commande.get("Aff_ID"))
        client_name = ""
        date_tir = ""
        id_client = ""
        if id_aff:
            affaire = graph_get_item_by_id(site_id, affaire_list_id, id_aff, token)
            if affaire:
                date_tir = fmt_date(affaire.get("Date_evt", ""))
                id_client = clean_id(affaire.get("Client_ID"))
                if id_client:
                    cli = graph_get_item_by_id(site_id, clients_list_id, id_client, token)
                    if cli:
                        client_name = cli.get("Title", "").replace("'", " ")

        # --- 3. Émetteur (fournisseur de l'entité) ---
        emetteur = ""
        try:
            fournisseurs = graph_filtered_items(
                site_id, admin_list_id, token,
                f"fields/Clef eq 'Fournisseur' and fields/Entite eq '{entite}'")
            if fournisseurs:
                f0 = fournisseurs[0]
                emetteur = (f"{f0.get('Title','')}\n{f0.get('Param_x00e8_tres','')}\n"
                            f"{f0.get('Parametres_4','')} {f0.get('Parametres_3','')}")
        except Exception as e:
            logging.warning(f"Fournisseur introuvable: {e}")

        # --- 4. Site de livraison ---
        ville_livraison = adresse_livraison = telephone = ""
        try:
            sites = graph_filtered_items(
                site_id, admin_list_id, token,
                f"fields/Clef eq 'Site_livraison' and fields/Title eq '{site_livraison_title}' and fields/Entite eq '{entite}'")
            if sites:
                s0 = sites[0]
                ville_livraison = f"{s0.get('Parametres_2','')} - {s0.get('Parametres_3','')}"
                adresse_livraison = s0.get("Param_x00e8_tres", "")
                telephone = s0.get("Parametres_4", "")
        except Exception as e:
            logging.warning(f"Site de livraison introuvable: {e}")

        # --- 5. Détails commande ---
        details = fetch_details(site_id, details_list_id, token, commande_id)
        logging.info(f"{len(details)} lignes de détail trouvées.")

        # --- 6. Produits : une seule map (par Title = Produit_ID), 1ère occurrence ---
        produits = graph_filtered_items(site_id, produits_list_id, token)
        produits_map = {}
        for p in produits:
            key = str(p.get("Title", "")).strip().upper()
            if key and key not in produits_map:
                produits_map[key] = p

        # --- 7. Matériel : réservations (par Aff_ID) + liste matériel (par Materiel_ID) ---
        reservations = []
        if id_aff:
            reservations = fetch_by_field(site_id, materiel_resa_list_id, token, "Aff_ID", id_aff)
        materiel_items = graph_filtered_items(site_id, materiel_list_id, token)
        materiel_map = {}
        for m in materiel_items:
            # Materiel_ID = champ interne `Title` (affiché « Materiel_ID »)
            key = clean_id(m.get("Title"))
            if key:
                materiel_map[key] = m

        # --- 8. Génération : un fichier par site de stock ---
        # Nettoyage des anciens éléments commande_doc (ID_cmd + Type_Doc)
        cleanup_old_docs(site_id, commande_doc_list_id, token, commande_id, type_doc)

        timestamp = datetime.utcnow().strftime('%Y-%m-%d-%H-%M-%S')
        template_bytes = download_graph_file(site_id, token, TEMPLATE_PATH)

        # Définition des sites de stock à générer
        stock_targets = []
        if site_stock:
            stock_targets.append({"name": site_stock, "is_secondary": False})
        if site_stock_second:
            stock_targets.append({"name": site_stock_second, "is_secondary": True})
        if not stock_targets:
            # Aucun site de stock défini : on génère tout de même un fichier "principal"
            stock_targets.append({"name": "", "is_secondary": False})

        inventaire_cache = {}
        generated = []

        for target in stock_targets:
            is_sec = target["is_secondary"]
            dest_name = target["name"]
            # Valeur utilisée pour cibler l'inventaire / filtrer les lignes
            site_filter = site_stock_second if site_stock_second else "sec"

            # Filtrage des détails selon le site de stock
            if is_sec:
                details_site = [d for d in details if str(d.get("Site", "")) == str(site_stock_second)]
            else:
                details_site = [d for d in details if str(d.get("Site", "")) != str(site_filter)]

            # Réservations matériel filtrées par site
            if is_sec:
                resa_site = [r for r in reservations if str(r.get("Site", "")) == str(site_stock_second)]
            else:
                resa_site = [r for r in reservations if str(r.get("Site", "")) != str(site_filter)]

            header_data = {
                "Emetteur": emetteur,
                "Destinataire": dest_name,
                "Commande N°": numero_commande,
                "Client": client_name,
                "Adresse Livraison": adresse_livraison,
                "Ville Livraison": ville_livraison,
                "Telephone": telephone,
                "Site de livraison": site_livraison_title,
                "Date de livraison": date_livraison,
                "Date de tir": date_tir,
                "CREATEUR": "",
            }

            fiche_rows = build_fiche_prepa_rows(
                details_site, produits_map, inventaire_cache,
                site_id, inventaire_list_id, details_list_id, token,
                site_filter, is_sec)
            accessoires_rows = build_accessoires_rows(details_site, produits_map)
            # Onglet Global = produits + accessoires (tout sauf UKOBA)
            grouped_rows = build_global_rows(fiche_rows, accessoires_rows)
            materiel_rows = build_materiel_rows(resa_site, materiel_map)

            file_bytes = fill_workbook(template_bytes, header_data, fiche_rows,
                                       grouped_rows, accessoires_rows, materiel_rows)

            # Nom de fichier
            suffix = f"_{dest_name}" if dest_name else ""
            nom_fichier = f"FP-{numero_commande}-{client_name}{suffix}_{timestamp}.xlsx"
            nom_fichier_doc = f"FP-{numero_commande}-{client_name}{suffix}"

            # Élément commande_doc
            item_id = graph_post_item(site_id, commande_doc_list_id, token, {
                "Title": nom_fichier_doc,
                "ID_cmd": str(commande_id),
                "ID_Aff": str(id_aff) if id_aff else "",
                "ID_Client": str(id_client) if id_client else "",
                "Type_Doc": type_doc,
            })

            # Upload du fichier dans la bibliothèque (comme GenerationDocument)
            drive_item_id, file_list_item_id = upload_file(site_id, token, nom_fichier, file_bytes)

            generated.append({
                "site_stock": dest_name,
                "filename": nom_fichier,
                "created_item_id": item_id,
                "drive_item_id": drive_item_id,
                "file_list_item_id": file_list_item_id,
                "nb_lignes": len(fiche_rows),
                "nb_accessoires": len(accessoires_rows),
                "nb_materiel": len(materiel_rows),
            })

        return func.HttpResponse(
            json.dumps({
                "status": "success",
                "message": f"{len(generated)} fiche(s) de préparation générée(s).",
                "Type_Doc": type_doc,
                "files": generated,
            }),
            status_code=200, mimetype="application/json")

    except Exception as e:
        logging.exception("Erreur dans GenerationFichePrepa")
        return _err(f"Erreur serveur interne : {str(e)}", 500)


# =========================================================================
#  HELPERS POINT D'ENTRÉE
# =========================================================================
def _err(message, code):
    return func.HttpResponse(
        json.dumps({"status": "error", "message": message}),
        status_code=code, mimetype="application/json")


def fetch_by_field(site_id, list_id, token, field, value):
    """Récupère des éléments par un champ donné (texte puis nombre)."""
    items = []
    try:
        items = graph_filtered_items(site_id, list_id, token, f"fields/{field} eq '{value}'")
    except Exception:
        pass
    if not items:
        try:
            items = graph_filtered_items(site_id, list_id, token, f"fields/{field} eq {value}")
        except Exception:
            pass
    return items


def cleanup_old_docs(site_id, list_id, token, commande_id, type_doc):
    try:
        headers = {"Authorization": f"Bearer {token}", "Prefer": "HonorNonIndexedQueriesWarningMayFailRandomly"}
        for filt in (f"fields/ID_cmd eq '{commande_id}' and fields/Type_Doc eq '{type_doc}'",
                     f"fields/ID_cmd eq {commande_id} and fields/Type_Doc eq '{type_doc}'"):
            fp = urllib.parse.quote(filt, safe="=()/ '")
            url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items?$expand=fields&$filter={fp}"
            res = session.get(url, headers=headers)
            if res.status_code == 400:
                continue
            if res.ok:
                for old in res.json().get("value", []):
                    oid = old.get("id")
                    if oid:
                        session.delete(f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items/{oid}",
                                       headers={"Authorization": f"Bearer {token}"})
                        logging.info(f"Ancien document supprimé: {oid}")
                break
    except Exception as e:
        logging.error(f"Erreur nettoyage anciens documents: {e}")


def upload_file(site_id, token, nom_fichier, file_bytes):
    """Upload le fichier généré dans la bibliothèque (même approche que GenerationDocument)."""
    url = (f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/"
           f"{TARGET_FOLDER}/{urllib.parse.quote(nom_fichier, safe='')}:/content")
    res = session.put(url, headers={"Authorization": f"Bearer {token}",
                                    "Content-Type": "application/octet-stream"}, data=file_bytes)
    res.raise_for_status()
    drive_item = res.json()
    drive_item_id = drive_item["id"]

    file_list_item_id = ""
    try:
        url_li = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{drive_item_id}?$expand=listItem"
        res_li = session.get(url_li, headers={"Authorization": f"Bearer {token}"})
        if res_li.ok:
            file_list_item_id = res_li.json().get("listItem", {}).get("id", "")
    except Exception as e:
        logging.warning(f"Impossible de récupérer le listItemId: {e}")

    return drive_item_id, file_list_item_id
