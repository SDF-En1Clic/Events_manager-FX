import logging
import azure.functions as func
from azure.identity import DefaultAzureCredential
from azure.keyvault.secrets import SecretClient
import requests
import urllib.parse
import json
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
import re

# --------- CONFIG GLOBALE -------------
VAULT_URL = "https://events-manager-kv.vault.azure.net/"
# --------------------------------------

session = requests.Session()
adapter = requests.adapters.HTTPAdapter(pool_connections=100, pool_maxsize=100)
session.mount('https://', adapter)

def get_secret(name: str):
    credential = DefaultAzureCredential()
    client = SecretClient(vault_url=VAULT_URL, credential=credential)
    return client.get_secret(name).value

def get_graph_token(tenant_id, client_id, client_secret):
    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default"
    }
    headers = { "Content-Type": "application/x-www-form-urlencoded" }
    try:
        response = session.post(url, data=data, headers=headers)
        response.raise_for_status()
        return response.json().get("access_token")
    except Exception as e:
        logging.error(f"Erreur token: {e}")
        return None

def graph_filtered_items(site_id, list_id, token, filter_expr=None):
    base_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items?$expand=fields"
    headers = {
        "Authorization": f"Bearer {token}",
        "Prefer": "HonorNonIndexedQueriesWarningMayFailRandomly"
    }
    if filter_expr:
        filter_param = urllib.parse.quote(filter_expr, safe="=()/ ")
        base_url += f"&$filter={filter_param}"
    
    results = []
    url = base_url
    while url:
        res = session.get(url, headers=headers)
        if not res.ok: 
            logging.error(f"Erreur API Graph (filtré). Status: {res.status_code}. Réponse: {res.text}")
        res.raise_for_status()
        data = res.json()
        results.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return [item.get("fields", {}) for item in results]

def graph_get_item_by_id(site_id, list_id, item_id, token):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items/{item_id}?$expand=fields"
    headers = { "Authorization": f"Bearer {token}" }
    try:
        res = session.get(url, headers=headers)
        res.raise_for_status()
        return res.json().get("fields", {})
    except Exception as e:
        logging.error(f"Erreur Get Item: {e}")
        return None

def download_graph_file(site_id, token, filepath):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    headers = { "Authorization": f"Bearer {token}" }
    res = session.get(url, headers=headers)
    res.raise_for_status()
    drives = res.json().get("value", [])
    
    drive_id = None
    for d in drives:
        if d.get("name") == "Documents partages" or d.get("name") == "Documents partagés":
            drive_id = d.get("id")
            break
            
    if not drive_id and drives:
        drive_id = drives[0].get("id")
        
    url_file = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{urllib.parse.quote(filepath)}:/content"
    res_file = session.get(url_file, headers=headers)
    res_file.raise_for_status()
    return res_file.content

def graph_post_item(site_id, list_id, token, fields_dict):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    payload = { "fields": fields_dict }
    res = session.post(url, headers=headers, json=payload)
    if not res.ok:
        logging.error(f"Erreur create item: {res.status_code} {res.text}")
    res.raise_for_status()
    return res.json().get("id")

def append_to_table(sheet, table, data_list):
    if not data_list: return
    
    match = re.search(r'([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)', table.ref)
    if not match: return
    
    start_col, start_row, end_col, end_row = match.groups()
    start_row = int(start_row)
    
    start_col_idx = column_index_from_string(start_col)
    end_col_idx = column_index_from_string(end_col)
    
    headers = []
    for col_idx in range(start_col_idx, end_col_idx + 1):
        headers.append(sheet.cell(row=start_row, column=col_idx).value)
        
    current_row = start_row
    for row_data in data_list:
        current_row += 1
        for col_idx, header in enumerate(headers, start=start_col_idx):
            # Utilise .get() : si la clé n'existe pas, écrit "" par défaut
            sheet.cell(row=current_row, column=col_idx, value=row_data.get(header, ""))
            
    table.ref = f"{start_col}{start_row}:{end_col}{current_row}"

# --- MODIFICATION : Ajout de flat_data et grouped_data en paramètres ---
def fill_excel(template_bytes, header_data, flat_data, grouped_data, is_ukoba, sheet_name=None):
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    
    table_DataL = None
    table_Datas = None
    table_Datas2 = None
    
    for sheet in wb.worksheets:
        for tb in sheet.tables.values():
            if tb.name == 'tabDataL':
                table_DataL = (tb, sheet)
            elif tb.name == 'tabDatas':
                table_Datas = (tb, sheet)
            elif tb.name == 'tabDatas2':
                table_Datas2 = (tb, sheet)
                
    if table_DataL:
        append_to_table(table_DataL[1], table_DataL[0], [header_data])
        
    if is_ukoba:
        # Ukoba : tabDatas reçoit la liste plate classique, tabDatas2 reçoit la liste groupée
        if table_Datas:
            append_to_table(table_Datas[1], table_Datas[0], flat_data)
        if table_Datas2:
            append_to_table(table_Datas2[1], table_Datas2[0], grouped_data)
    else:
        # Commande Globale : tabDatas reçoit la liste groupée
        if table_Datas:
            append_to_table(table_Datas[1], table_Datas[0], grouped_data)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def main(req: func.HttpRequest) -> func.HttpResponse:
    try:
        body = req.get_json()
        commande_id = body.get("ID_cmd")
        type_doc = body.get("Type_Doc")
        
        logging.info(f"Paramètre reçus : ID_cmd={commande_id}, Type_Doc={type_doc}")
        if not commande_id or not type_doc:
            return func.HttpResponse(json.dumps({"status": "error", "message": "Paramètres 'ID_cmd' et 'Type_Doc' requis"}), status_code=400, mimetype="application/json")
            
        if type_doc not in ["Commande globale", "Commande ukoba", "Plan de tir"]:
            return func.HttpResponse(json.dumps({"status": "error", "message": f"Type_Doc non supporté: {type_doc}"}), status_code=400, mimetype="application/json")

        tenant_id = get_secret("tenantid")
        client_id = get_secret("clientid")
        client_secret = get_secret("appsecret")
        site_id = get_secret("siteid")
        
        commandes_list_id = get_secret("cmdlistid")
        details_list_id = get_secret("cmddetailslistid")
        produits_list_id = get_secret("produitslistid")
        admin_list_id = get_secret("configlistid")
        affaire_list_id = get_secret("affaireevtslistid")
        clients_list_id = get_secret("clientslistid")
        commande_doc_list_id = get_secret("commandedoclistid")

        token = get_graph_token(tenant_id, client_id, client_secret)
        if not token:
            return func.HttpResponse(json.dumps({"status": "error", "message": "Échec de l'authentification Graph"}), status_code=500, mimetype="application/json")

        # 1. Infos Commande 
        commande = graph_get_item_by_id(site_id, commandes_list_id, commande_id, token)
        if not commande:
            return func.HttpResponse(json.dumps({"status": "error", "message": "Commande introuvable"}), status_code=404, mimetype="application/json")
        
        entite = commande.get("Entite", "")
        numero_commande = commande.get("Title", "")
        site_livraison_title = commande.get("Site_livraison", "")
        date_livraison = commande.get("Date_livraison", "")
        if date_livraison:
            try:
                date_livraison = datetime.strptime(date_livraison[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                pass
        
        # 2. Infos Affaire et Client
        def clean_id(val):
            if not val: return ""
            if isinstance(val, float):
                return str(int(val))
            return str(val)

        id_aff = clean_id(commande.get("Aff_ID"))
        client_name = ""
        date_tir = ""
        id_client = ""
        type_vente = ""
        if id_aff:
            affaire = graph_get_item_by_id(site_id, affaire_list_id, id_aff, token)
            if affaire:
                date_tir = affaire.get("Date_evt", "")
                if date_tir:
                    try:
                        date_tir = datetime.strptime(date_tir[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
                    except Exception:
                        pass
                type_vente = affaire.get("Type_vente", "")
                id_client = clean_id(affaire.get("Client_ID"))
                if id_client:
                    client_meta = graph_get_item_by_id(site_id, clients_list_id, id_client, token)
                    if client_meta:
                        client_name = client_meta.get("Title", "").replace("'", " ")

        # 3. Récupération Emetteur/Destinataire
        fournisseurs = graph_filtered_items(site_id, admin_list_id, token, "fields/Clef eq 'Fournisseur'")
        emetteur = ""
        destinataire = ""
        for f in fournisseurs:
            title = f.get("Title", "")
            if title == "SOIRS DE FETES" and entite == "BASSIN PARISIEN":
                emetteur = f"{title}\n{f.get('Param_x00e8_tres','')}\n{f.get('Parametres_2','')} {f.get('Parametres_3','')}"
            elif title == "SOIRS DE FETES GRAND SUD" and entite == "GRAND SUD":
                emetteur = f"{title}\n{f.get('Param_x00e8_tres','')}\n{f.get('Parametres_2','')} {f.get('Parametres_3','')}"
            elif title == "UKOBA":
                destinataire = f"{title}\n{f.get('Param_x00e8_tres','')}\n{f.get('Parametres_2','')} {f.get('Parametres_3','')}"
        
        # 4. Informations Site Livraison
        sites = graph_filtered_items(site_id, admin_list_id, token, f"fields/Clef eq 'Site_livraison' and fields/Title eq '{site_livraison_title}' and fields/Entite eq '{entite}'")
        ville_livraison = ""
        adresse_livraison = ""
        telephone = ""
        if sites:
            site_info = sites[0]
            ville_livraison = f"{site_info.get('Parametres_2','')} - {site_info.get('Parametres_3','')}"
            adresse_livraison = site_info.get("Param_x00e8_tres","")
            telephone = site_info.get("Parametres_4","")

        # 5. Récupération des Détails et Produits
        logging.info(f"--- RECHERCHE DES LIGNES POUR CMD_ID : {commande_id} ---")
        
        details = []
        
        # 1. On essaie d'abord avec le format Texte ('1771')
        try:
            details = graph_filtered_items(site_id, details_list_id, token, f"fields/CMD_ID eq '{commande_id}'")
            if details:
                logging.info(f"Succès avec le format Texte : {len(details)} lignes trouvées.")
        except Exception:
            pass

        # 2. SI (et seulement si) on n'a rien trouvé, on essaie avec le format Nombre (1771)
        if not details:
            try:
                details = graph_filtered_items(site_id, details_list_id, token, f"fields/CMD_ID eq {commande_id}")
                if details:
                    logging.info(f"Succès avec le format Nombre : {len(details)} lignes trouvées.")
            except Exception:
                pass
            
        logging.info(f"Total des lignes prêtes pour Excel : {len(details)}")
        
        # Mapping des produits
        produits = graph_filtered_items(site_id, produits_list_id, token)
        produits_map = {}
        for p in produits:
            ref_brute = str(p.get("Title", "")).strip().upper()
            if ref_brute:
                produits_map[ref_brute] = p
            
        header_data = {
            "Emetteur": emetteur,
            "Destinataire": destinataire,
            "Commande N°": numero_commande,
            "Client": client_name,
            "Adresse Livraison": adresse_livraison,
            "Ville Livraison": ville_livraison,
            "Telephone": telephone,
            "Site de livraison": site_livraison_title,
            "Date de livraison": date_livraison,
            "Date de tir": date_tir
        }
        
        is_ukoba = (type_doc == "Commande ukoba")

        # --- TRI ET GROUPEMENT ADAPTÉS SELON LE TYPE DE DOC ---
        if is_ukoba:
            # Pour Ukoba : on trie par Référence (ordre alphabétique)
            details.sort(key=lambda x: str(x.get("Reference", "")).strip().upper())
        else:
            # Pour Globale : on trie par Ligne (Tri intelligent avec les nombres en premier)
            def tri_lignes(item):
                val = str(item.get("Title", "")).replace("Ligne", "").strip()
                try:
                    return (0, float(val))
                except ValueError:
                    return (1, val)
            details.sort(key=tri_lignes)

        # Création des deux listes
        flat_data = []     # Tableau avec des lignes classiques répétées
        grouped_data = []  # Tableau avec En-tête de groupe + enfants
        
        current_group = None

        for det in details:
            ref = str(det.get("Reference", "")).strip().upper()
            prod = produits_map.get(ref, {})
            origine = prod.get("Origine", "")
            
            if is_ukoba and origine != "UKOBA":
                continue
                
            ligne_titre = f"Ligne {det.get('Title', '')}"
            designation = prod.get("Description_pyromotion", "")
            qte = det.get("Quantite", "")
            commentaire = det.get("Commentaires", "")
            site_val = det.get("Site", "")
            statut = det.get("Statut", "")
            
            prix = ""
            if not is_ukoba:
                p_grossiste = prod.get("Prix_vente_grossiste_HT")
                p_presta = prod.get("Prix_vente_presta_HT")
                prix = p_grossiste if type_vente == "Produit" else p_presta

 # --- FILTRAGE PRÉLIMINAIRE ---
        filtered_details = []
        for det in details:
            ref = str(det.get("Reference", "")).strip().upper()
            prod = produits_map.get(ref, {})
            origine = prod.get("Origine", "")
            # Si document Ukoba, on ignore ce qui n'est pas Ukoba
            if is_ukoba and origine != "UKOBA":
                continue
            filtered_details.append(det)

        # --- FONCTION DE TRI PAR LIGNE ---
        def tri_lignes(item):
            val = str(item.get("Title", "")).replace("Ligne", "").strip()
            try:
                return (0, float(val))
            except ValueError:
                return (1, val)

        flat_data = []
        grouped_data = []

        # =========================================================
        # 1. FORMAT PLAT (Tableau 1 Ukoba) - Trié par LIGNE
        # =========================================================
        if is_ukoba:
            filtered_details.sort(key=tri_lignes)
            for det in filtered_details:
                ref = str(det.get("Reference", "")).strip().upper()
                prod = produits_map.get(ref, {})
                ligne_titre = f"Ligne {det.get('Title', '')}"
                designation = prod.get("Description_pyromotion", "")
                qte = det.get("Quantite", "")
                commentaire = det.get("Commentaires", "")
                
                flat_data.append({
                    "LIGNE": ligne_titre,
                    "REFERENCE": ref,
                    "QT": qte,
                    "DESIGNATION": designation,
                    "COMMENTAIRE": commentaire
                })

        # =========================================================
        # 2. FORMAT GROUPÉ (Tableau 2 Ukoba & Globale)
        # =========================================================
        if is_ukoba:
            # --- LOGIQUE UKOBA : GROUPÉ PAR RÉFÉRENCE ---
            # On trie d'abord la liste par Référence (Ordre alphabétique)
            filtered_details.sort(key=lambda x: str(x.get("Reference", "")).strip().upper())
            current_ref = None
            
            for det in filtered_details:
                ref = str(det.get("Reference", "")).strip().upper()
                prod = produits_map.get(ref, {})
                ligne_titre = f"Ligne {det.get('Title', '')}"
                designation = prod.get("Description_pyromotion", "")
                qte = det.get("Quantite", "")
                commentaire = det.get("Commentaires", "")
                
                # Rupture de groupe (Nouvelle Référence)
                if ref != current_ref:
                    current_ref = ref
                    # LIGNE D'EN-TÊTE : Ne contient QUE la référence
                    grouped_data.append({
                        "REFERENCE": ref
                    })
                
                # LIGNE ENFANT : Contient le reste (Designation, Qté, Ligne...), mais REFERENCE est vide
                grouped_data.append({
                    "LIGNE": ligne_titre,
                    "REFERENCE": "", 
                    "DESIGNATION": designation,
                    "QT": qte,
                    "COMMENTAIRE": commentaire
                })
                
        else:
            # --- LOGIQUE GLOBALE : GROUPÉ PAR LIGNE DE TIR ---
            # On trie la liste par numéro de Ligne
            filtered_details.sort(key=tri_lignes)
            current_ligne = None
            
            for det in filtered_details:
                ref = str(det.get("Reference", "")).strip().upper()
                prod = produits_map.get(ref, {})
                ligne_titre = f"Ligne {det.get('Title', '')}"
                designation = prod.get("Description_pyromotion", "")
                qte = det.get("Quantite", "")
                site_val = det.get("Site", "")
                statut = det.get("Statut", "")
                origine = prod.get("Origine", "")
                
                p_grossiste = prod.get("Prix_vente_grossiste_HT")
                p_presta = prod.get("Prix_vente_presta_HT")
                prix = p_grossiste if type_vente == "Produit" else p_presta
                
                # Rupture de groupe (Nouvelle Ligne de tir)
                if ligne_titre != current_ligne:
                    current_ligne = ligne_titre
                    # LIGNE D'EN-TÊTE : Ne contient QUE le nom de la ligne
                    grouped_data.append({
                        "LIGNE": ligne_titre
                    })
                
                # LIGNE ENFANT : Contient les infos du produit, mais LIGNE est vide
                grouped_data.append({
                    "LIGNE": "",
                    "REFERENCE": ref,
                    "DESIGNATION": designation,
                    "QT": qte,
                    "SITE": site_val,
                    "STATUT": statut,
                    "ORIGINE": origine,
                    "PRIX DE VENTE UNITAIRE HT": prix
                })

        # 6. Génération Fichier (EXCEL ou CSV)
        timestampStr = datetime.utcnow().strftime('%Y-%m-%d-%H-%M-%S')
        
        if type_doc == "Plan de tir":
            csv_lines = []
            for det in details:
                ref = str(det.get("Reference", "")).strip().upper()
                prod = produits_map.get(ref, {})
                if prod.get("Origine", "") == "UKOBA":
                    ligne = f"Ligne {det.get('Title', '')}"
                    qte_raw = det.get("Quantite", "")
                    try:
                        f_qte = float(qte_raw)
                        qte = str(int(f_qte)) if f_qte.is_integer() else str(f_qte)
                    except Exception:
                        qte = str(qte_raw)
                    csv_lines.append(f"{ligne};{ref};{qte}")
            
            csv_content = "\n".join(csv_lines) + "\n" if csv_lines else ""
            file_bytes = csv_content.encode('utf-8-sig')
            nom_fichier = f"SDF-{numero_commande}_Plan_de_tir_{timestampStr}.csv"
            nom_fichier_doc = f"SDF-{numero_commande}_Plan_de_tir"
            
        else:
            if is_ukoba:
                template_path = "00 - Templates/Template_commande_Ukoba_V1.xlsx"
                nom_fichier = f"SDF-{numero_commande}_{site_livraison_title}_{timestampStr}.xlsx"
                nom_fichier_doc = f"SDF-{numero_commande}_{site_livraison_title}.xlsx"
            else:
                template_path = "00 - Templates/Template_commande_globale_V1.xlsx"
                nom_fichier = f"SDF-{numero_commande}-{client_name}_{timestampStr}.xlsx"
                nom_fichier_doc = f"SDF-{numero_commande}-{client_name}.xlsx"

            template_bytes = download_graph_file(site_id, token, template_path)
            file_bytes = fill_excel(template_bytes, header_data, flat_data, grouped_data, is_ukoba)
        
        # 6.5 Suppression des anciens éléments avec le même nom (nettoyage)
        try:
            filter_param = urllib.parse.quote(f"fields/Title eq '{nom_fichier_doc}'", safe="=()/ ")
            url_check = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{commande_doc_list_id}/items?$expand=fields&$filter={filter_param}"
            headers_check = {"Authorization": f"Bearer {token}", "Prefer": "HonorNonIndexedQueriesWarningMayFailRandomly"}
            res_check = session.get(url_check, headers=headers_check)
            if res_check.ok:
                for old_item in res_check.json().get("value", []):
                    old_id = old_item.get("id")
                    if old_id:
                        url_del = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{commande_doc_list_id}/items/{old_id}"
                        session.delete(url_del, headers={"Authorization": f"Bearer {token}"})
                        logging.info(f"Ancien document SharePoint supprimé: {old_id}")
        except Exception as e:
            logging.error(f"Erreur lors de la suppression de l'ancien document: {e}")

        # 7. Création de l'élément dans SharePoint (Liste Commande Document)
        new_item_fields = {
            "Title": nom_fichier_doc,
            "ID_cmd": str(commande_id),
            "ID_Aff": str(id_aff) if id_aff else "",
            "ID_Client": str(id_client) if id_client else "",
            "Type_Doc": type_doc
        }
        item_id = graph_post_item(site_id, commande_doc_list_id, token, new_item_fields)
        
        # 8. Upload du document généré temporairement dans "00 - Templates"
        url_upload = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/03 - Documents commandes/{urllib.parse.quote(nom_fichier, safe='')}:/content"
        headers_upload = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream"
        }
        res_upload = session.put(url_upload, headers=headers_upload, data=file_bytes)
        res_upload.raise_for_status()
        drive_item = res_upload.json()
        drive_item_id = drive_item['id']
        
        return func.HttpResponse(
            json.dumps({
                "status": "success",
                "message": f"Le document {type_doc} a été créé avec succès.",
                "created_item_id": item_id, 
                "item_id": item_id,
                "Type_Doc": type_doc,
                "drive_item_id": drive_item_id, 
                "filename": nom_fichier
            }),
            status_code=200,
            mimetype="application/json"
        )
        
    except Exception as e:
        logging.exception("Erreur dans GenerationDocument")
        return func.HttpResponse(
            json.dumps({"status": "error", "message": f"Erreur serveur interne : {str(e)}"}), 
            status_code=500, 
            mimetype="application/json"
        )